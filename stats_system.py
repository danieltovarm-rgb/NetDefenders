"""
Sistema de puntuaciones y registro de errores para NetDefenders
Implementación con POO
"""

import time
import json
import os
from datetime import datetime
from typing import List, Dict, Optional
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ========== SISTEMA DE PUNTUACIONES POR NIVEL ==========

class ScoreManager:
    """
    Maneja el sistema de puntuaciones del juego por nivel.
    """
    def __init__(self, level_number=1):
        self.level_number = level_number
        self.current_score = 0
        self.best_score = 0  # mejor puntaje histórico del nivel
        self.combo = 0
        self.score_multiplier = 1.0
    
    def add_points(self, points, with_combo=True):
        """Añade puntos, considera combo si está activo"""
        if with_combo and self.combo > 0:
            bonus = min(self.combo * 0.1, 2.0)  # max 2x
            points = int(points * (1 + bonus))
        
        self.current_score += int(points * self.score_multiplier)
        self.combo += 1
        
        # actualizar mejor puntaje si superamos el anterior
        if self.current_score > self.best_score:
            self.best_score = self.current_score
        
    def subtract_points(self, penalty):
        """Resta puntos por error"""
        self.current_score = max(0, self.current_score - penalty)
        self.combo = 0
    
    def reset_combo(self):
        """Resetea la racha"""
        self.combo = 0
    
    def get_rank(self, score=None):
        """Devuelve ranking basado en puntos (usa current_score si no se especifica)"""
        score_to_rank = score if score is not None else self.current_score
        if score_to_rank >= 5000:
            return "S"
        elif score_to_rank >= 3000:
            return "A"
        elif score_to_rank >= 1500:
            return "B"
        elif score_to_rank >= 500:
            return "C"
        else:
            return "D"
    
    def reset_current_score(self):
        """Reinicia puntuación actual (mantiene best_score)"""
        self.current_score = 0
        self.combo = 0


# ========== SISTEMA DE REGISTRO DE ERRORES ==========

class Mistake:
    """Representa un error/equivocación individual"""
    def __init__(self, level: int, mistake_type: str, description: str, 
                 mistake_details: Dict = None, timestamp: str = None):
        self.level = level
        self.type = mistake_type
        self.description = description
        self.mistake_details = mistake_details or {}  # Detalles específicos: logo, dominio, texto
        self.timestamp = timestamp or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.severity = self._calculate_severity()
    
    def _calculate_severity(self):
        """Calcula gravedad: baja, media, alta, crítica"""
        critical_types = ["contraseña_expuesta", "fuga_datos", "malware_ejecutado"]
        high_types = ["malware_no_detectado", "descarga_sospechosa", "phishing_no_detectado"]
        medium_types = ["falso_positivo", "respuesta_lenta"]
        
        if self.type in critical_types:
            return "crítica"
        elif self.type in high_types:
            return "alta"
        elif self.type in medium_types:
            return "media"
        else:
            return "baja"
    
    def get_severity_color(self):
        """Retorna color RGB según gravedad"""
        colors = {
            "crítica": (220, 20, 20),
            "alta": (255, 100, 0),
            "media": (255, 200, 0),
            "baja": (100, 200, 100)
        }
        return colors.get(self.severity, (255, 255, 255))
    
    def to_dict(self):
        """Convierte a diccionario para guardar/mostrar"""
        return {
            "nivel": self.level,
            "tipo": self.type,
            "descripcion": self.description,
            "detalles": self.mistake_details,
            "fecha_hora": self.timestamp,
            "gravedad": self.severity
        }


class MistakeLog:
    """
    Tabla/log de errores. Almacena todas las equivocaciones.
    Solo guarda datos de la PRIMERA vez que se juega cada nivel.
    Exporta a Excel o JSON.
    """
    def __init__(self, excel_filename: str = "datos_recolectados.xlsx"):
        self.excel_filename = excel_filename
        self.mistakes: List[Mistake] = []
        self.mistake_counts: Dict[str, int] = {}
        self.first_play_per_level: Dict[int, bool] = {}  # Track si es primera vez en cada nivel
        self._load_existing_data()
    
    def _load_existing_data(self):
        """Carga datos existentes para saber qué niveles ya se jugaron"""
        if not EXCEL_AVAILABLE or not os.path.exists(self.excel_filename):
            return
        
        try:
            wb = openpyxl.load_workbook(self.excel_filename)
            if "Registro_Errores" in wb.sheetnames:
                ws = wb["Registro_Errores"]
                # Identificar niveles ya registrados
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Si hay nivel
                        self.first_play_per_level[row[0]] = False  # Ya se jugó
            wb.close()
        except Exception as e:
            print(f"Error cargando datos existentes: {e}")
    
    def is_first_play(self, level: int) -> bool:
        """Verifica si es la primera vez jugando este nivel"""
        return self.first_play_per_level.get(level, True)
    
    def mark_level_played(self, level: int):
        """Marca un nivel como ya jugado"""
        self.first_play_per_level[level] = False
    
    def add_mistake(self, level: int, mistake_type: str, description: str, 
                   mistake_details: Dict = None, force_save: bool = False):
        """Registra una nueva equivocación (solo en primera jugada o si force_save=True)"""
        # Solo guardar si es primera vez jugando este nivel o force_save
        if not self.is_first_play(level) and not force_save:
            return  # No guardar en jugadas posteriores
        
        mistake = Mistake(level, mistake_type, description, mistake_details)
        self.mistakes.append(mistake)
        
        # actualizar contador
        self.mistake_counts[mistake_type] = self.mistake_counts.get(mistake_type, 0) + 1
    
    def get_mistakes_by_level(self, level: int) -> List[Mistake]:
        """Obtiene errores de un nivel específico"""
        return [m for m in self.mistakes if m.level == level]
    
    def get_total_mistakes(self) -> int:
        """Total de errores"""
        return len(self.mistakes)
    
    def save_to_excel(self):
        """Guarda el log en Excel con formato"""
        if not EXCEL_AVAILABLE:
            print("openpyxl no está instalado. Usando JSON como alternativa.")
            return self.save_to_json()
        
        try:
            # Crear o cargar workbook
            if os.path.exists(self.excel_filename):
                wb = openpyxl.load_workbook(self.excel_filename)
            else:
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
            
            # Crear o seleccionar hoja
            if "Registro_Errores" not in wb.sheetnames:
                ws = wb.create_sheet("Registro_Errores")
                # Crear encabezados
                headers = ["Nivel", "Fecha/Hora", "Tipo de Acción", "Descripción", 
                          "Error en Logo", "Error en Dominio", "Error en Texto", "Número de Errores"]
                ws.append(headers)
                
                # Estilo de encabezados
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                ws = wb["Registro_Errores"]
            
            # Agregar todas las acciones con contador de errores individual por acción
            action_count = len([row for row in ws.iter_rows(min_row=2)]) + 1  # Contar acciones existentes
            
            for mistake in self.mistakes:
                details = mistake.mistake_details
                
                # Contar errores en esta acción específica (Logo, Dominio, Texto marcados incorrectamente)
                errores_en_esta_accion = 0
                if details.get("logo"):  # Si cometió error en Logo
                    errores_en_esta_accion += 1
                if details.get("dominio"):  # Si cometió error en Dominio
                    errores_en_esta_accion += 1
                if details.get("texto"):  # Si cometió error en Texto
                    errores_en_esta_accion += 1
                
                # Si no hay errores específicos pero la acción es un error general, contar como 1
                if errores_en_esta_accion == 0 and mistake.type in ["phishing_no_detectado", "falso_positivo"]:
                    errores_en_esta_accion = 1
                
                ws.append([
                    mistake.level,
                    mistake.timestamp,
                    mistake.type,
                    mistake.description,
                    "Sí" if details.get("logo") else "No",
                    "Sí" if details.get("dominio") else "No",
                    "Sí" if details.get("texto") else "No",
                    errores_en_esta_accion  # Puede ser 0, 1, 2, o 3 dependiendo de cuántos errores
                ])
                action_count += 1
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar
            wb.save(self.excel_filename)
            wb.close()
            
            # Limpiar lista temporal (ya se guardó)
            self.mistakes.clear()
            
            return True
        except Exception as e:
            print(f"Error guardando Excel: {e}")
            return self.save_to_json()
    
    def save_to_json(self):
        """Guarda en JSON como alternativa"""
        filename = self.excel_filename.replace('.xlsx', '.json')
        try:
            # Cargar datos existentes si existen
            existing_data = []
            if os.path.exists(filename):
                with open(filename, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            
            # Agregar nuevos errores
            for mistake in self.mistakes:
                existing_data.append(mistake.to_dict())
            
            # Guardar todo
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(existing_data, f, indent=2, ensure_ascii=False)
            
            # Limpiar lista temporal
            self.mistakes.clear()
            
            return True
        except Exception as e:
            print(f"Error guardando JSON: {e}")
            return False
    
    def clear(self):
        """Limpia el log temporal (no afecta archivo guardado)"""
        self.mistakes.clear()
        self.mistake_counts.clear()


# ========== ESTADÍSTICAS DEL JUGADOR POR NIVEL ==========

class PlayerStats:
    """
    Centraliza estadísticas por nivel.
    - Puntajes separados por nivel
    - Solo guarda errores en primera jugada
    - Actualiza mejor puntaje si se supera
    """
    def __init__(self, player_name: str = "Jugador"):
        self.name = player_name
        
        # Sistema de puntuación por nivel (3 niveles)
        self.level_scores = {
            1: ScoreManager(level_number=1),
            2: ScoreManager(level_number=2),
            3: ScoreManager(level_number=3)
        }
        
        # Sistema de registro de errores (único para todos los niveles)
        self.mistake_log = MistakeLog()
        
        # Nivel actual
        self.current_level = 1
        
        # Estadísticas temporales por sesión (no se guardan permanentemente)
        self.session_stats = {
            "emails_analyzed": 0,
            "threats_detected": 0,
            "threats_missed": 0,
            "false_positives": 0
        }
        # Niveles completados (éxito) para lógica de quiz final
        self.completed_levels = set()
        # === QUIZ (Nivel 1+2 combinado) ===
        self.pre_quiz_score = 0
        self.post_quiz_score = 0
        self.quiz_total_questions = 12  # fijo (6 phishing + 6 malware)
        self.quiz_improvement = 0.0
        # Breakdown por categoría
        self.pre_quiz_level1_correct = 0
        self.pre_quiz_level2_correct = 0
        self.post_quiz_level1_correct = 0
        self.post_quiz_level2_correct = 0
    
    def set_current_level(self, level: int):
        """Cambia el nivel actual"""
        if level in self.level_scores:
            self.current_level = level
            # Resetear puntuación actual del nivel para nueva partida
            self.level_scores[level].reset_current_score()
    
    def get_current_score_manager(self) -> ScoreManager:
        """Obtiene el ScoreManager del nivel actual"""
        return self.level_scores.get(self.current_level)
    
    def analyze_email(self, is_threat: bool, detected_correctly: bool, 
                     mistake_details: Dict = None):
        """
        Registra análisis de un email.
        Guarda TODAS las acciones (aciertos y errores) en Excel/JSON si es primera vez jugando el nivel.
        """
        self.session_stats["emails_analyzed"] += 1
        score_mgr = self.get_current_score_manager()
        
        if is_threat and detected_correctly:
            # Detectó correctamente una amenaza
            self.session_stats["threats_detected"] += 1
            score_mgr.add_points(100)
            
            # NUEVO: Registrar acierto también
            self.mistake_log.add_mistake(
                level=self.current_level,
                mistake_type="amenaza_detectada_correctamente",
                description="Amenaza identificada y reportada correctamente",
                mistake_details=mistake_details or {}
            )
            
        elif is_threat and not detected_correctly:
            # Dejó pasar una amenaza (ERROR GRAVE)
            self.session_stats["threats_missed"] += 1
            
            # Guardar error si es primera jugada del nivel
            self.mistake_log.add_mistake(
                level=self.current_level,
                mistake_type="phishing_no_detectado",
                description="Amenaza no detectada en el correo",
                mistake_details=mistake_details or {}
            )
            
            score_mgr.subtract_points(80)
            
        elif not is_threat and detected_correctly:
            # Identificó correctamente un email legítimo
            score_mgr.add_points(50)
            
            # NUEVO: Registrar acierto también
            self.mistake_log.add_mistake(
                level=self.current_level,
                mistake_type="correo_legitimo_identificado",
                description="Correo legítimo procesado correctamente",
                mistake_details=mistake_details or {}
            )
            
        else:  # not is_threat and not detected_correctly
            # Falso positivo (marcó como amenaza algo legítimo)
            self.session_stats["false_positives"] += 1
            
            # Guardar error si es primera jugada del nivel
            self.mistake_log.add_mistake(
                level=self.current_level,
                mistake_type="falso_positivo",
                description="Correo legítimo marcado como amenaza",
                mistake_details=mistake_details or {}
            )
            
            score_mgr.subtract_points(30)
    
    def complete_level(self):
        """
        Completa el nivel actual.
        Guarda errores a Excel si es primera jugada.
        Actualiza mejor puntaje si se superó.
        """
        score_mgr = self.get_current_score_manager()
        
        # Actualizar mejor puntaje si se superó
        if score_mgr.current_score > score_mgr.best_score:
            score_mgr.best_score = score_mgr.current_score
        
        # Guardar errores a Excel solo si fue primera jugada
        if self.mistake_log.is_first_play(self.current_level):
            self.mistake_log.save_to_excel()
            self.mistake_log.mark_level_played(self.current_level)
        
        # Limpiar errores temporales
        self.mistake_log.clear()
        # Marcar nivel completado (éxito)
        self.completed_levels.add(self.current_level)
    
    def get_accuracy(self) -> float:
        """Calcula precisión de la sesión actual (0-100)"""
        total = self.session_stats["emails_analyzed"]
        if total == 0:
            return 100.0
        
        correct = (self.session_stats["threats_detected"] + 
                  (total - self.session_stats["threats_detected"] - 
                   self.session_stats["threats_missed"] - 
                   self.session_stats["false_positives"]))
        return (correct / total) * 100
    
    def get_level_rank(self, level: int) -> str:
        """Obtiene el rango del nivel basado en su mejor puntaje"""
        if level in self.level_scores:
            return self.level_scores[level].get_rank(self.level_scores[level].best_score)
        return "D"
    
    def get_level_best_score(self, level: int) -> int:
        """Obtiene el mejor puntaje de un nivel"""
        if level in self.level_scores:
            return self.level_scores[level].best_score
        return 0
    
    def get_level_current_score(self, level: int) -> int:
        """Obtiene el puntaje actual de un nivel"""
        if level in self.level_scores:
            return self.level_scores[level].current_score
        return 0
    
    def reset_session_stats(self):
        """Reinicia estadísticas de la sesión"""
        self.session_stats = {
            "emails_analyzed": 0,
            "threats_detected": 0,
            "threats_missed": 0,
            "false_positives": 0
        }

    # ================== QUIZ MÉTODOS ==================
    def record_quiz_score(self, mode: str, total_correct: int, cat_breakdown: dict):
        """Registra resultado del quiz.
        mode: 'pre' | 'post' | 'final' (final se trata como post)
        cat_breakdown: { 'level1': int, 'level2': int }
        """
        if mode == 'pre':
            self.pre_quiz_score = total_correct
            self.pre_quiz_level1_correct = cat_breakdown.get('level1', 0)
            self.pre_quiz_level2_correct = cat_breakdown.get('level2', 0)
        elif mode in ('post','final'):
            self.post_quiz_score = total_correct
            self.post_quiz_level1_correct = cat_breakdown.get('level1', 0)
            self.post_quiz_level2_correct = cat_breakdown.get('level2', 0)
            # Calcular mejora porcentual respecto a total de preguntas
            if self.quiz_total_questions > 0:
                pre_pct = (self.pre_quiz_score / self.quiz_total_questions) * 100
                post_pct = (self.post_quiz_score / self.quiz_total_questions) * 100
                self.quiz_improvement = post_pct - pre_pct
            else:
                self.quiz_improvement = 0.0
            # Registrar en mistake_log para telemetría agregada
            self.mistake_log.add_mistake(
                level=self.current_level,
                mistake_type="quiz_mejora",
                description="Resultado quiz final registrado" if mode=='final' else "Resultado post-quiz registrado",
                mistake_details={
                    "pre_total": self.pre_quiz_score,
                    "post_total": self.post_quiz_score,
                    "mejora_pct": round(self.quiz_improvement, 2),
                    "pre_lvl1": self.pre_quiz_level1_correct,
                    "pre_lvl2": self.pre_quiz_level2_correct,
                    "post_lvl1": self.post_quiz_level1_correct,
                    "post_lvl2": self.post_quiz_level2_correct
                }
            )

    def get_quiz_summary(self) -> dict:
        return {
            "pre_total": self.pre_quiz_score,
            "post_total": self.post_quiz_score,
            "mejora_pct": round(self.quiz_improvement, 2),
            "pre_lvl1": self.pre_quiz_level1_correct,
            "pre_lvl2": self.pre_quiz_level2_correct,
            "post_lvl1": self.post_quiz_level1_correct,
            "post_lvl2": self.post_quiz_level2_correct
        }


# ========== SISTEMA DE NIVEL 2: RECURSOS Y LIMPIEZA DE PC ==========

class ResourceBar:
    """
    Gestiona la barra de recursos del jugador en Nivel 2.
    Los recursos disminuyen con acciones y síntomas activos.
    """
    def __init__(self, initial_resources=100, max_resources=100):
        self.current = initial_resources
        self.max = max_resources
        self.passive_drain_rate = 0  # recursos/segundo por síntomas
        
    def consume(self, amount):
        """Consume recursos por una acción"""
        self.current = max(0, self.current - amount)
        return self.current > 0
    
    def passive_drain(self, dt):
        """Drenaje pasivo por síntomas activos (dt en milisegundos)"""
        if self.passive_drain_rate > 0:
            drain = (self.passive_drain_rate * dt) / 1000.0
            self.current = max(0, self.current - drain)
    
    def restore(self, amount):
        """Restaura recursos (por bonus o logros)"""
        self.current = min(self.max, self.current + amount)
    
    def set_drain_rate(self, rate):
        """Establece la tasa de drenaje pasivo"""
        self.passive_drain_rate = max(0, rate)
    
    def is_depleted(self):
        """Verifica si los recursos se agotaron"""
        return self.current <= 0
    
    def get_percentage(self):
        """Retorna el porcentaje de recursos restantes"""
        return (self.current / self.max) * 100 if self.max > 0 else 0


class Symptom:
    """
    Representa un síntoma visible causado por un virus.
    """
    def __init__(self, symptom_type, severity, resource_drain):
        self.type = symptom_type  # "ralentizacion", "popups", "pantalla_bloqueada", "teclas_fantasma"
        self.severity = severity  # 1-10
        self.resource_drain = resource_drain  # recursos/segundo
        self.active = False
        self.source_file = None  # referencia al archivo que lo causa
        
    def activate(self, source_file=None):
        """Activa el síntoma"""
        self.active = True
        self.source_file = source_file
    
    def deactivate(self):
        """Desactiva el síntoma"""
        self.active = False
        self.source_file = None


class SymptomManager:
    """
    Gestiona todos los síntomas activos y su impacto.
    """
    def __init__(self):
        self.symptoms = {
            "ralentizacion": Symptom("ralentizacion", 6, 2.0),
            "popups": Symptom("popups", 5, 1.5),
            "pantalla_bloqueada": Symptom("pantalla_bloqueada", 10, 5.0),
            "teclas_fantasma": Symptom("teclas_fantasma", 4, 1.0)
        }
    
    def activate_symptom(self, symptom_type, source_file=None):
        """Activa un síntoma"""
        if symptom_type in self.symptoms:
            self.symptoms[symptom_type].activate(source_file)
            return True
        return False
    
    def deactivate_symptom(self, symptom_type):
        """Desactiva un síntoma"""
        if symptom_type in self.symptoms:
            self.symptoms[symptom_type].deactivate()
            return True
        return False
    
    def get_total_drain(self):
        """Calcula el drenaje total de recursos por síntomas activos"""
        total = 0
        for symptom in self.symptoms.values():
            if symptom.active:
                total += symptom.resource_drain
        return total
    
    def get_active_symptoms(self):
        """Retorna lista de síntomas activos"""
        return [s for s in self.symptoms.values() if s.active]
    
    def has_active_symptoms(self):
        """Verifica si hay síntomas activos"""
        return any(s.active for s in self.symptoms.values())


class ActionTimer:
    """
    Temporizador para acciones con duración (escaneo, limpieza, etc.)
    """
    def __init__(self, action_name, duration, cost):
        self.action_name = action_name
        self.duration = duration  # milisegundos
        self.cost = cost  # costo en recursos
        self.elapsed = 0
        self.in_progress = False
        self.completed = False
        
    def start(self):
        """Inicia el temporizador"""
        self.in_progress = True
        self.elapsed = 0
        self.completed = False
    
    def update(self, dt):
        """Actualiza el temporizador"""
        if self.in_progress:
            self.elapsed += dt
            if self.elapsed >= self.duration:
                self.in_progress = False
                self.completed = True
                return True
        return False
    
    def reset(self):
        """Reinicia el temporizador"""
        self.elapsed = 0
        self.in_progress = False
        self.completed = False
    
    def get_progress(self):
        """Retorna el progreso de la acción (0-1)"""
        if self.duration <= 0:
            return 1.0
        return min(1.0, self.elapsed / self.duration)


class ActionTimerFactory:
    """
    Factory para crear temporizadores de acciones con configuración balanceada
    """
    ACTION_CONFIGS = {
        "inspeccionar": {"duration": 500, "cost": 0},
        "escanear_archivo": {"duration": 3000, "cost": 10},
        "escanear_carpeta": {"duration": 5000, "cost": 15},
        "cuarentena": {"duration": 2000, "cost": 8},
        "limpiar_malware": {"duration": 4000, "cost": 0},
        "limpiar_seguro": {"duration": 1500, "cost": 12}
    }
    
    @classmethod
    def create(cls, action_name):
        """Crea un temporizador para la acción especificada"""
        config = cls.ACTION_CONFIGS.get(action_name, {"duration": 1000, "cost": 5})
        return ActionTimer(action_name, config["duration"], config["cost"])


class VictoryConditionChecker:
    """
    Verifica las condiciones de victoria del Nivel 2
    """
    def __init__(self):
        self.total_threats = 0
        self.threats_eliminated = 0
        self.threats_quarantined = 0
        
    def set_total_threats(self, count):
        """Establece el total de amenazas en el nivel"""
        self.total_threats = count
    
    def register_elimination(self):
        """Registra una amenaza eliminada"""
        self.threats_eliminated += 1
    
    def register_quarantine(self):
        """Registra una amenaza en cuarentena"""
        self.threats_quarantined += 1
    
    def check_victory(self):
        """Verifica si se cumplieron las condiciones de victoria"""
        total_handled = self.threats_eliminated + self.threats_quarantined
        return total_handled >= self.total_threats and self.total_threats > 0
    
    def get_completion_percentage(self):
        """Retorna el porcentaje de amenazas manejadas"""
        if self.total_threats == 0:
            return 0
        total_handled = self.threats_eliminated + self.threats_quarantined
        return (total_handled / self.total_threats) * 100


class DefeatConditionChecker:
    """
    Verifica las condiciones de derrota del Nivel 2
    """
    def __init__(self, resource_bar):
        self.resource_bar = resource_bar
        self.critical_threshold = 10  # recursos críticos
        
    def check_defeat(self):
        """Verifica si se cumplió la condición de derrota"""
        return self.resource_bar.is_depleted()
    
    def is_critical(self):
        """Verifica si los recursos están en nivel crítico"""
        return self.resource_bar.current <= self.critical_threshold


class QuizBonusSystem:
    """
    Sistema de bonus por completar el quiz final
    """
    def __init__(self):
        self.quiz_completed = False
        self.correct_answers = 0
        self.total_questions = 0
        self.bonus_points = 0
        
    def complete_quiz(self, correct, total):
        """Registra los resultados del quiz"""
        self.quiz_completed = True
        self.correct_answers = correct
        self.total_questions = total
        self.bonus_points = self._calculate_bonus()
    
    def _calculate_bonus(self):
        """Calcula los puntos de bonus según el rendimiento"""
        if self.total_questions == 0:
            return 0
        
        percentage = (self.correct_answers / self.total_questions) * 100
        
        if percentage >= 90:
            return 500  # Excelente
        elif percentage >= 70:
            return 300  # Bueno
        elif percentage >= 50:
            return 150  # Regular
        else:
            return 50   # Bajo
    
    def get_bonus(self):
        """Retorna los puntos de bonus"""
        return self.bonus_points


class Level2GameManager:
    """
    Gestiona el flujo completo del Nivel 2: recursos, síntomas, victoria/derrota
    """
    def __init__(self, total_threats=0):
        self.resource_bar = ResourceBar(initial_resources=100, max_resources=100)
        self.symptom_manager = SymptomManager()
        self.victory_checker = VictoryConditionChecker()
        self.defeat_checker = DefeatConditionChecker(self.resource_bar)
        self.quiz_bonus = QuizBonusSystem()
        self.score_manager = ScoreManager(level_number=2)
        
        self.game_state = "playing"  # "playing", "victory", "defeat"
        self.start_time = time.time()
        
        if total_threats > 0:
            self.victory_checker.set_total_threats(total_threats)
    
    def update(self, dt):
        """Actualiza el estado del juego (dt en milisegundos)"""
        if self.game_state != "playing":
            return
        
        # Aplicar drenaje pasivo de síntomas
        total_drain = self.symptom_manager.get_total_drain()
        self.resource_bar.set_drain_rate(total_drain)
        self.resource_bar.passive_drain(dt)
        
        # Verificar condiciones de derrota
        if self.defeat_checker.check_defeat():
            self.game_state = "defeat"
            return
        
        # Verificar condiciones de victoria
        if self.victory_checker.check_victory():
            self.game_state = "victory"
            self._apply_victory_bonus()
    
    def execute_action(self, action_name):
        """Ejecuta una acción y consume recursos"""
        timer = ActionTimerFactory.create(action_name)
        
        # Verificar si hay recursos suficientes
        if not self.resource_bar.consume(timer.cost):
            return None  # No hay recursos suficientes
        
        return timer
    
    def file_cleaned(self, had_virus, symptom_type=None):
        """Registra que un archivo fue limpiado"""
        if had_virus:
            self.victory_checker.register_elimination()
            self.score_manager.add_points(100, with_combo=True)
            
            # Desactivar síntoma asociado
            if symptom_type:
                self.symptom_manager.deactivate_symptom(symptom_type)
        else:
            # Penalización por limpiar archivo seguro
            self.score_manager.subtract_points(50)
    
    def file_quarantined(self, had_virus, symptom_type=None):
        """Registra que un archivo fue puesto en cuarentena"""
        if had_virus:
            self.victory_checker.register_quarantine()
            self.score_manager.add_points(80, with_combo=True)
            
            # Desactivar síntoma asociado
            if symptom_type:
                self.symptom_manager.deactivate_symptom(symptom_type)
        else:
            # Penalización menor por cuarentena innecesaria
            self.score_manager.subtract_points(30)
    
    def file_scanned(self, is_infected):
        """Registra que un archivo fue escaneado"""
        if is_infected:
            self.score_manager.add_points(20, with_combo=False)
    
    def activate_virus_symptom(self, symptom_type, source_file=None):
        """Activa un síntoma de virus"""
        self.symptom_manager.activate_symptom(symptom_type, source_file)
    
    def complete_quiz(self, correct, total):
        """Completa el quiz y aplica bonus"""
        self.quiz_bonus.complete_quiz(correct, total)
        bonus = self.quiz_bonus.get_bonus()
        self.score_manager.add_points(bonus, with_combo=False)
    
    def _apply_victory_bonus(self):
        """Aplica bonus por victoria"""
        # Bonus por recursos restantes
        resource_bonus = int(self.resource_bar.current * 5)
        self.score_manager.add_points(resource_bonus, with_combo=False)
        
        # Bonus por tiempo
        elapsed = time.time() - self.start_time
        if elapsed < 300:  # Menos de 5 minutos
            time_bonus = 500
        elif elapsed < 600:  # Menos de 10 minutos
            time_bonus = 300
        else:
            time_bonus = 100
        
        self.score_manager.add_points(time_bonus, with_combo=False)
    
    def get_game_stats(self):
        """Retorna estadísticas del juego"""
        return {
            "state": self.game_state,
            "resources": self.resource_bar.current,
            "resources_percentage": self.resource_bar.get_percentage(),
            "score": self.score_manager.current_score,
            "threats_handled": self.victory_checker.threats_eliminated + self.victory_checker.threats_quarantined,
            "total_threats": self.victory_checker.total_threats,
            "completion": self.victory_checker.get_completion_percentage(),
            "active_symptoms": len(self.symptom_manager.get_active_symptoms()),
            "is_critical": self.defeat_checker.is_critical(),
            "elapsed_time": time.time() - self.start_time
        }
